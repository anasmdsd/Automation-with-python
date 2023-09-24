import openpyxl as xl
def process_workbook():
    filename=input("enter the path:" )
    wb=xl.load_workbook(filename)
    sheet=wb['Sheet1']
    z=sheet.cell(1,1)
    print(z.value)
    print(sheet.max_row,'\n')

    for i in range(2, sheet.max_row+1):
        f=sheet.cell(i,3)
        v=f.value
        corrected_price=v*0.9
        new_cell=sheet.cell(i,4)#defining the row & column to store the data
        new_cell.value=corrected_price#storing the value
    New_price=sheet.cell(1,4)
    New_price.value="Correct_price"
    wb.save(filename)
process_workbook()
